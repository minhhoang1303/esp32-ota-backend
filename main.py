from flask import Flask, request, jsonify, send_file, session, redirect, url_for
from flask_socketio import SocketIO, emit
from werkzeug.utils import secure_filename
import os
import json
import time
import socket
import threading
from datetime import datetime, timedelta

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key')
app.config['UPLOAD_FOLDER'] = 'upload'
socketio = SocketIO(app, cors_allowed_origins="*")

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

registered_devices = {}
firmware_updates = {}
last_heartbeat_time = {}
last_commands = {}
device_sensor_data = {}
device_chart_data = {}
heartbeat_timeout = 20000

valid_username = os.environ.get('ADMIN_USER', 'minhhoangcdt')
valid_password = os.environ.get('ADMIN_PASS', '13032001Jr@')


def get_server_ip_address():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip_address = s.getsockname()[0]
        s.close()
        return ip_address
    except Exception:
        try:
            hostname = socket.gethostname()
            ip_address = socket.gethostbyname(hostname)
            return ip_address
        except Exception:
            return '127.0.0.1'


def get_server_port():
    return 1313


def load_registered_devices_from_file():
    file_path = 'registered_devices.txt'
    try:
        if os.path.exists(file_path):
            with open(file_path, 'r') as f:
                devices = json.load(f)
                for device in devices:
                    registered_devices[device['hostName']] = device
            print('Registered devices loaded successfully.')
        else:
            print('No registered devices file found.')
    except Exception as e:
        print(f'Error loading registered devices: {e}')


def save_registered_devices_to_file():
    file_path = 'registered_devices.txt'
    try:
        with open(file_path, 'w') as f:
            json.dump(list(registered_devices.values()), f, indent=2)
        print('Registered devices saved successfully.')
    except Exception as e:
        print(f'Error saving registered devices: {e}')


@socketio.on('connect')
def handle_connect():
    print('Client connected')


@socketio.on('disconnect')
def handle_disconnect():
    print('Client disconnected')


@socketio.on('message')
def handle_message(message):
    print(f'Received message from client: {message}')
    socketio.emit('message', message, skip_sid=request.sid)


@app.route('/')
def login_page():
    return send_file('login.html')


@app.route('/index.html')
def index_page():
    if session.get('isLoggedIn'):
        return send_file('index.html')
    return redirect(url_for('login_page'))


@app.route('/login', methods=['POST'])
def login():
    data = request.get_json()
    if not data:
        data = request.form

    username = data.get('username')
    password = data.get('password')

    print(f'Received login request: {username}, {password}')

    if username == valid_username and password == valid_password:
        session['isLoggedIn'] = True
        if request.is_json:
            return jsonify({'success': True, 'redirect': '/index.html'})
        else:
            return redirect(url_for('index_page'))
    else:
        if request.is_json:
            return jsonify({'success': False, 'error': 'Invalid username or password'}), 401
        else:
            return 'Invalid username or password.', 401


@app.route('/authenticate', methods=['POST'])
def authenticate():
    data = request.get_json()
    if not data:
        data = request.form

    password = data.get('password')

    if password == valid_password:
        return '', 200
    else:
        return '', 401


@app.route('/logout', methods=['POST'])
def logout():
    session['isLoggedIn'] = False
    return '', 200


@app.route('/firmwareInitiated', methods=['POST'])
def firmware_initiated():
    host_name = request.args.get('hostName')
    print(f'Received "Hello" message from {host_name}')
    socketio.emit('message', f'Received Firmware OTA by {host_name}')
    return '', 200


@app.route('/getWebSocketAddress', methods=['GET'])
def get_websocket_address():
    server_ip = get_server_ip_address()
    server_port = get_server_port()

    return jsonify({
        'ip': server_ip,
        'port': server_port
    })


@app.route('/register', methods=['POST'])
def register():
    data = request.get_data(as_text=True).strip().split('\n')

    if len(data) >= 5:
        host_name = data[0].strip()
        firmware_version = data[1].strip()
        mac_address = data[2].strip()
        try:
            wifi_signal_strength = int(data[3].strip())
        except ValueError:
            wifi_signal_strength = 0
        ip_address = data[4].strip()

        if host_name and firmware_version and mac_address and ip_address:
            if host_name in registered_devices:
                if registered_devices[host_name]['firmwareVersion'] != firmware_version:
                    print(f'Firmware version changed for {host_name}. Re-registering...')
                    firmware_updates[host_name] = True
            else:
                print(f'Registering {host_name}')

            registered_devices[host_name] = {
                'hostName': host_name,
                'firmwareVersion': firmware_version,
                'macAddress': mac_address,
                'wifiSignalStrength': wifi_signal_strength,
                'ipAddress': ip_address,
                'password': registered_devices[host_name].get('password', '') if host_name in registered_devices else ''
            }

            print('Registered devices:', list(registered_devices.keys()))
            save_registered_devices_to_file()
            return 'Registration successful.', 200

    return 'Bad Request.', 400


@app.route('/getDevices', methods=['GET'])
def get_devices():
    return jsonify(list(registered_devices.keys()))


@app.route('/getFirmwareVersion', methods=['GET'])
def get_firmware_version():
    host_name = request.args.get('hostName')
    if not host_name:
        return 'Bad Request.', 400

    device_info = registered_devices.get(host_name)
    if device_info:
        return device_info['firmwareVersion'], 200
    else:
        return 'Firmware version not found.', 404


@app.route('/upload', methods=['POST'])
def upload_firmware():
    host_name = request.args.get('hostName')
    if not host_name:
        return 'Bad Request.', 400

    if 'firmwareFile' not in request.files:
        return 'No file uploaded.', 400

    file = request.files['firmwareFile']
    if file.filename == '':
        return 'No file selected.', 400

    filename = secure_filename(f"{host_name}_firmware.bin")
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(file_path)

    print(f'Received firmware binary for {host_name}.')

    firmware_updates[host_name] = True

    return 'Firmware upload successful.', 200


@app.route('/updateStatus', methods=['GET'])
def update_status():
    host_name = request.args.get('hostName')
    if not host_name:
        return 'Bad Request.', 400

    update_available = firmware_updates.get(host_name, False)
    if update_available:
        firmware_updates[host_name] = False
        return 'Update Available', 200
    else:
        return 'No Update Available', 204


@app.route('/heartbeat', methods=['POST'])
def heartbeat():
    host_name = request.get_data(as_text=True)
    if not host_name:
        return 'Bad Request.', 400

    last_heartbeat_time[host_name] = time.time() * 1000
    return f'Heartbeat received from {host_name}.', 200


@app.route('/flushAllDevices', methods=['POST'])
def flush_all_devices():
    registered_devices.clear()
    last_heartbeat_time.clear()
    save_registered_devices_to_file()
    return '', 200


@app.route('/getOnlineStatus', methods=['GET'])
def get_online_status():
    device_status_list = []
    now = time.time() * 1000

    for device, device_info in registered_devices.items():
        last_heartbeat = last_heartbeat_time.get(device)
        online = last_heartbeat and (now - last_heartbeat <= heartbeat_timeout)

        device_status_list.append({
            'device': device,
            'online': online,
            'firmwareVersion': device_info['firmwareVersion'],
            'macAddress': device_info['macAddress'],
            'wifiSignalStrength': device_info['wifiSignalStrength'],
            'ipAddress': device_info['ipAddress'],
            'hasPassword': bool(device_info.get('password', ''))
        })

    return jsonify(device_status_list)


@app.route('/upload/<path:filename>', methods=['GET'])
def download_firmware(filename):
    filename = secure_filename(filename)
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    if os.path.exists(file_path) and os.path.isfile(file_path):
        return send_file(file_path, as_attachment=True, download_name=filename)
    else:
        return 'Firmware not found.', 404


@app.route('/sendSerialData', methods=['POST'])
def send_serial_data():
    host_name = request.args.get('hostName')
    serial_data = request.get_data(as_text=True)

    print(f'Received serial data from {host_name}: {serial_data}')
    return 'Serial data received successfully', 200


@app.route('/sendCommand', methods=['POST'])
def send_command():
    data = request.get_json()
    if not data:
        return 'Bad Request.', 400

    device_host_name = data.get('hostName')
    command = data.get('command')

    if not device_host_name:
        return 'Bad Request.', 400

    if command:
        last_commands[device_host_name] = command
        print(f'Received command "{command}" from ESP8266 for {device_host_name}')
    else:
        print(f'Sending command "{last_commands.get(device_host_name)}" to {device_host_name}')

    return last_commands.get(device_host_name, ''), 200


@app.route('/getCommand', methods=['GET'])
def get_command():
    device_host_name = request.args.get('hostName')
    command = last_commands.get(device_host_name, '')

    if device_host_name in last_commands:
        del last_commands[device_host_name]

    return command, 200


@app.route('/getServerVersion', methods=['GET'])
def get_server_version():
    return "1.0.3", 200


@app.route('/dashboard')
def dashboard_page():
    if session.get('isLoggedIn'):
        return send_file('dashboard.html')
    return redirect(url_for('login_page'))


@app.route('/device/<host_name>', methods=['GET'])
def device_dashboard(host_name):
    if host_name not in registered_devices:
        return 'Device not found.', 404
    
    device_info = registered_devices[host_name]
    password = device_info.get('password', '')
    
    if password:
        provided_password = request.args.get('password', '')
        if provided_password != password:
            return jsonify({'error': 'Unauthorized'}), 401
    
    now = time.time() * 1000
    last_heartbeat = last_heartbeat_time.get(host_name)
    online = last_heartbeat and (now - last_heartbeat <= heartbeat_timeout)
    
    sensor_history = device_sensor_data.get(host_name, [])
    latest_sensor = sensor_history[-1] if sensor_history else None
    
    pending_command = last_commands.get(host_name, '')
    
    return jsonify({
        'hostName': host_name,
        'firmwareVersion': device_info['firmwareVersion'],
        'macAddress': device_info['macAddress'],
        'wifiSignalStrength': device_info['wifiSignalStrength'],
        'ipAddress': device_info['ipAddress'],
        'online': online,
        'lastHeartbeat': last_heartbeat,
        'latestSensor': latest_sensor,
        'pendingCommand': pending_command
    })


@app.route('/sensorData', methods=['POST'])
def sensor_data():
    raw_data = request.get_data(as_text=True).replace('\r\n', '\n').replace('\r', '\n').strip()
    data = raw_data.split('\n')
    
    if len(data) >= 3:
        host_name = data[0].strip()
        try:
            temperature = float(data[1].strip())
            humidity = float(data[2].strip())
        except ValueError:
            return 'Bad Request: Invalid sensor data.', 400
        
        current_time = time.time()
        
        if host_name not in device_sensor_data:
            device_sensor_data[host_name] = []
        
        device_sensor_data[host_name].append({
            'timestamp': current_time,
            'temperature': temperature,
            'humidity': humidity
        })
        
        if len(device_sensor_data[host_name]) > 50:
            device_sensor_data[host_name] = device_sensor_data[host_name][-50:]
        
        localtime = time.localtime(current_time)
        minute = localtime.tm_min
        if minute % 10 == 0:
            if host_name not in device_chart_data:
                device_chart_data[host_name] = []
            
            device_chart_data[host_name].append({
                'timestamp': current_time,
                'temperature': temperature,
                'humidity': humidity
            })
            
            if len(device_chart_data[host_name]) > 50:
                device_chart_data[host_name] = device_chart_data[host_name][-50:]
        
        print(f'[SENSOR] {host_name}: Temp={temperature}C, Humidity={humidity}%')
        return 'Sensor data received.', 200
    
    return 'Bad Request.', 400


@app.route('/getSensorData/<host_name>', methods=['GET'])
def get_sensor_data(host_name):
    limit = int(request.args.get('limit', 50))
    sensor_data = device_sensor_data.get(host_name, [])
    return jsonify(sensor_data[-limit:])


@app.route('/getChartData/<host_name>', methods=['GET'])
def get_chart_data(host_name):
    limit = int(request.args.get('limit', 50))
    chart_data = device_chart_data.get(host_name, [])
    return jsonify(chart_data[-limit:])


@app.route('/getLatestSensor/<host_name>', methods=['GET'])
def get_latest_sensor(host_name):
    sensor_data = device_sensor_data.get(host_name, [])
    if sensor_data:
        return jsonify(sensor_data[-1])
    return jsonify(None)


@app.route('/clearChartData', methods=['POST'])
def clear_chart_data():
    data = request.get_json()
    device = data.get('device')
    password = data.get('password', '')
    
    device_info = registered_devices.get(device)
    if not device_info:
        return jsonify({'success': False, 'message': 'Thiết bị không tồn tại'}), 404
    
    if device_info.get('password'):
        if password != device_info['password']:
            return jsonify({'success': False, 'message': 'Sai mật khẩu'}), 401
    
    if device and device in device_chart_data:
        device_chart_data[device] = []
        return jsonify({'success': True})
    return jsonify({'success': False}), 400


@app.route('/downloadExcel/<host_name>', methods=['GET'])
def download_excel(host_name):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    
    chart_data = device_chart_data.get(host_name, [])
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Du lieu cam bien"
    
    header = ['STT', 'Thời gian', 'Nhiệt độ (°C)', 'Độ ẩm (%)']
    for col, h in enumerate(header, 1):
        ws.cell(row=1, column=col, value=h)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for i, d in enumerate(chart_data, 2):
        ws.cell(row=i, column=1, value=i-1)
        ws.cell(row=i, column=2, value=datetime.fromtimestamp(d['timestamp']).strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=i, column=3, value=round(d['temperature'], 1))
        ws.cell(row=i, column=4, value=round(d['humidity'], 1))
        
        for col in range(1, 5):
            ws.cell(row=i, column=col).border = thin_border
    
    for col in range(1, 5):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=1, column=col).border = thin_border
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2
    
    filepath = f'temp/sensor_{host_name}.xlsx'
    os.makedirs('temp', exist_ok=True)
    wb.save(filepath)
    
    return send_file(filepath, as_attachment=True, download_name=f'log_nhiet_do_{host_name}.xlsx')


@app.route('/getDashboardData', methods=['GET'])
def get_dashboard_data():
    target_host = request.args.get('hostName')
    now = time.time() * 1000
    
    devices_data = []
    devices_to_show = {target_host: registered_devices[target_host]} if target_host and target_host in registered_devices else registered_devices
    
    for device, device_info in devices_to_show.items():
        last_heartbeat = last_heartbeat_time.get(device)
        online = last_heartbeat and (now - last_heartbeat <= heartbeat_timeout)
        
        sensor_history = device_sensor_data.get(device, [])
        latest_sensor = sensor_history[-1] if sensor_history else None
        
        devices_data.append({
            'hostName': device,
            'firmwareVersion': device_info['firmwareVersion'],
            'macAddress': device_info['macAddress'],
            'wifiSignalStrength': device_info['wifiSignalStrength'],
            'ipAddress': device_info['ipAddress'],
            'online': online,
            'lastHeartbeat': last_heartbeat,
            'latestSensor': latest_sensor
        })
    
    all_sensor_data = []
    if target_host and target_host in device_sensor_data:
        for data in device_sensor_data[target_host][-20:]:
            all_sensor_data.append({
                'device': target_host,
                'timestamp': data.get('timestamp'),
                'temperature': data.get('temperature'),
                'humidity': data.get('humidity')
            })
    else:
        for device, history in device_sensor_data.items():
            for data in history[-20:]:
                all_sensor_data.append({
                    'device': device,
                    'timestamp': data.get('timestamp'),
                    'temperature': data.get('temperature'),
                    'humidity': data.get('humidity')
                })
    
    all_sensor_data.sort(key=lambda x: x['timestamp'])
    
    total_devices = len(devices_to_show)
    online_count = sum(1 for d in devices_data if d['online'])
    offline_count = total_devices - online_count
    
    avg_temp = 0
    avg_humidity = 0
    if all_sensor_data:
        avg_temp = sum(d['temperature'] for d in all_sensor_data) / len(all_sensor_data)
        avg_humidity = sum(d['humidity'] for d in all_sensor_data) / len(all_sensor_data)
    
    return jsonify({
        'totalDevices': total_devices,
        'onlineDevices': online_count,
        'offlineDevices': offline_count,
        'averageTemperature': round(avg_temp, 2),
        'averageHumidity': round(avg_humidity, 2),
        'devices': devices_data,
        'recentSensorData': all_sensor_data[-100:]
    })


def check_offline_devices():
    while True:
        try:
            now = time.time() * 1000
            devices_to_remove = []

            for device, heartbeat_time in last_heartbeat_time.items():
                if now - heartbeat_time > heartbeat_timeout:
                    devices_to_remove.append(device)

            for device in devices_to_remove:
                del last_heartbeat_time[device]

            time.sleep(1)
        except Exception as e:
            print(f"Error in offline device checker: {e}")
            time.sleep(1)


threading.Thread(target=check_offline_devices, daemon=True).start()

load_registered_devices_from_file()

if __name__ == '__main__':
    port = 1313
    print(f"Server starting on port {port}")
    print(f"Server IP address: {get_server_ip_address()}")
    app.run(host='0.0.0.0', port=port, debug=False, use_reloader=False)